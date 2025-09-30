import time
import requests
import openpyxl
from datetime import datetime
from typing import List, Dict
import json

# ==================================================
# == Stanford GPT Setup (Minimal Example)         ==
# ==================================================
RATE_LIMIT_TOKENS_PER_MINUTE_CHAT = 250000

#API_ENDPOINT_CHAT = "https://apim.stanfordhealthcare.org/openai3/deployments/gpt-4o/chat/completions?api-version=2023-05-15"
API_ENDPOINT_CHAT = "https://apim.stanfordhealthcare.org/openai-eastus2/deployments/o1/chat/completions?api-version=2024-12-01-preview"

SUBSCRIPTION_KEY_CHAT = ""


tokens_used_chat = 0

start_time_chat = time.time()


class MinimalStanfordSecureGPT:
    def __init__(self, temperature=0.1, max_tokens=4000):
        self.temperature = temperature
        self.max_tokens = max_tokens

    def invoke(self, prompt: str) -> str:
        """
        A basic way to call GPT-4o endpoint with exponential backoff,
        while respecting a 2,500 tokens/min rate limit.
        """
        global tokens_used_chat, start_time_chat
        max_retries = 5
        retries = 0
        backoff_factor = 2
        initial_wait = 1

        while retries < max_retries:
            # Calculate how many tokens we need. This includes the prompt + potential response.
            tokens_needed = (len(prompt) // 4) + self.max_tokens
            elapsed_time = time.time() - start_time_chat

            # If adding these tokens exceeds our per-minute limit, wait until the minute resets.
            if tokens_used_chat + tokens_needed > RATE_LIMIT_TOKENS_PER_MINUTE_CHAT:
                wait_time = 60 - elapsed_time
                if wait_time > 0:
                    print(f"Chat rate limit reached. Waiting {wait_time:.2f} seconds...")
                    time.sleep(wait_time)
                tokens_used_chat = 0
                start_time_chat = time.time()

            headers = {
                "Ocp-Apim-Subscription-Key": SUBSCRIPTION_KEY_CHAT,
                "Content-Type": "application/json",
            }
            # payload = {
            #     "messages": [{"role": "user", "content": prompt}],
            #     "temperature": self.temperature,
            #     "max_tokens": self.max_tokens,
            # }
            payload = json.dumps({
                "model": "o1",
                "messages": [
                    {
                    "role": "user",
                    "content": prompt
                    }
                ]
                })

            try:
                response = requests.request("POST", API_ENDPOINT_CHAT, headers=headers, data=payload)
                

                if response.status_code == 429:  # Too Many Requests
                    retry_after = int(response.headers.get("Retry-After", 0))
                    wait_time = (
                        retry_after
                        if retry_after > 0
                        else initial_wait * (backoff_factor ** retries)
                    )
                    print(f"429 Too Many Requests. Sleeping {wait_time}s...")
                    time.sleep(wait_time)
                    retries += 1
                    continue

                response.raise_for_status()
                result = response.json()

                # Update usage
                tokens_used_chat += tokens_needed
                return result["choices"][0]["message"]["content"]

            except requests.exceptions.HTTPError as http_err:
                print(f"HTTP error occurred: {http_err}")
                if retries < max_retries - 1:
                    wait_time = initial_wait * (backoff_factor ** retries)
                    print(f"Retrying in {wait_time} seconds...")
                    time.sleep(wait_time)
                    retries += 1
                else:
                    raise

            except requests.exceptions.Timeout:
                print("A timeout occurred while waiting for response. Retrying...")
                retries += 1
                wait_time = initial_wait * (backoff_factor ** retries)
                print(f"Retrying in {wait_time} seconds...")
                time.sleep(wait_time)
                continue

            except Exception as e:
                print(f"An error occurred: {e}")
                if retries < max_retries - 1:
                    wait_time = initial_wait * (backoff_factor ** retries)
                    print(f"Retrying in {wait_time} seconds...")
                    time.sleep(wait_time)
                    retries += 1
                else:
                    raise

        raise Exception(f"Failed to get response after {max_retries} retries")


EXAMPLE_COMPARISONS = '''Here are some examples of the human-generated comparisons for different variables:

GEOGRAPHY:
• Evenly divded between urban and rural settings
• Serving areas populated by multiple races and ethnicities, especially Hispanic, but also Black/African American, Asian, Caribbean, and by migrant/seasonal workers; often underserved, Medicaid, with high social needs burden
• Areas range from highly competitive, with multiple FQHCs in the area, to lacking needed specialty services
• Several FQHCs collaborate with community-based organizations and local government entities

PATIENTS:
• Widely diverse demographics in terms of race and ethnicity across clinics and often within individual clinics' patient populations (one noted its patients spoke 50 languages)
• Diabetes was almost invariably among the most common health conditions of FQHC patients; other chronic illnesses, including hypertension and asthma, as well as substance use and behavioral health conditions
• Predominantly low income or very low income, often resulting in substantial social needs including, most frequently transportation, but also housing, employment, food insecturity; one clinic specialized in care of homeless people
• Some patients in some FQHCs lacked reliable internet, due to access issues or patient preference
• One clinic reported lack of trust as a barrier to care for some patients

SUPPORT FOR PATIENT GOALS, NEEDS, AND PREFERENCES:
• FQHCs identify patient needs, goals, and prefrences for diabetes management through structured assessments and provider-patient interactions; not all clinics have formal systems, relying instead on strong patient-provider relationships
• Social workers, health educators, nutritionists, community health workers frequently use motivational interviewing and trauma-informed approaches to encourage pursuit of self-identified goals for self-care; these visits often combine social/behavioral health needs with diabetes care
• FQHCs often strive to offer culturally and linguistically competent care, including by hiring aligned clinicians and staff
• FQHCs deploy multiple creative strategies to enable patients to address social needs, e.g., offering longer appointments to build trust with providers, facilitating prior authorizations, co-locating specialist and screening services, physical spaces for conducting telehealth visits'''

def process_spreadsheet(input_xlsx: str, output_xlsx: str):
    """
    1) Load an Excel workbook where:
       - Column A = Variable
       - Column B = Variable description
       - Columns C.. = data from different sites
    2) For each row, build a GPT prompt that compares how each site handles that variable
       based on the variable description.
    3) Write the GPT comparison into a new column at the end.
    4) Save the workbook as we go.
    """

    print(f"[process_spreadsheet()] Loading workbook: {input_xlsx}")
    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb.active  # or choose a specific sheet by name

    # Prepare a GPT instance
    gpt = MinimalStanfordSecureGPT(temperature=0.0, max_tokens=4000)

    # Determine how many columns we currently have
    max_col = ws.max_column
    max_row = ws.max_row

    # Insert a header in the last+1 column for "Comparison across Sites"
    comparison_col_index = max_col + 1
    ws.insert_cols(idx=comparison_col_index)
    ws.cell(row=1, column=comparison_col_index).value = "Comparison"

    # Get site names from the first row
    site_names = {}
    for col_idx in range(4, max_col + 1):
        site_name = ws.cell(row=1, column=col_idx).value
        if site_name:
            site_names[col_idx] = site_name
    
    # Loop through each row, skipping the header row
    for row_idx in range(2, 26):
        # Only process specific rows
        # if row_idx not in [6, 10, 15, 18, 23]:
        #     continue
            
        variable = ws.cell(row=row_idx, column=1).value
        if "technology for clinical care" not in variable.lower():
            continue
        else:
            print("found")
        
        var_description = ws.cell(row=row_idx, column=2).value

        print(f"\n{'='*80}")
        print(f"Processing Row {row_idx}")
        print(f"Variable: {variable}")
        print(f"Description: {var_description}")
        print('='*80)

        # If variable or description is missing, skip
        if not variable or not var_description:
            print("Skipping row - missing variable or description")
            continue

        # Gather site data from columns C.. up to max_col
        site_data_texts = []
        for col_idx in range(4, max_col + 1):
            site_value = ws.cell(row=row_idx, column=col_idx).value
            if site_value:
                site_name = site_names.get(col_idx, f"Site {col_idx-2}")  # Fallback to Site X if name not found
                site_data_texts.append(f"{site_name}: {site_value}")
                print(f"{site_name} data: {site_value}")

        # If there's no data from any site, skip
        if not site_data_texts:
            print("Skipping row - no site data available")
            continue

        # Build a comparison prompt for GPT

    
        prompt = (
            f"You are comparing data from different federally qualified health centers.\n\n"
            f"**Variable:** {variable}\n"
            f"**Variable Description:** {var_description}\n\n"
            f"Below are notes from different sites:\n"
        )
        for sd in site_data_texts:
            prompt += f"- {sd}\n"
        
        # Build the prompt using a different approach to avoid syntax errors
        prompt += "Write a summary of the data from the different sites in bullet points with the same format, \n"
        prompt += "length, and style as the human-generated comparisons. By 'create a summary', I mean come up with key themes and write them up, \n"
        prompt += "as if you were creating the actual text to be included in a cross-site summary document given to sites. The goal is summarizing what sites are doing, \n"
        prompt += "while highlighting the best of what they are doing; we want to be concise and provide actionable insights. \n"
        prompt += "Here are some examples of the human-generated comparisons for different variables:\n\n"
        
        prompt += "Example 1: \n"
        prompt += "Variable: Organization \n"
        prompt += "Variable description: Salient features of the FQHC, including organization size, infrastructure, i.e., number/types of clinics, services offered, major/ongoing organizational changes (e.g., mergers, re-structuring, expansions, new facilities), operations and financial arrangements including payment/compensation model and payer mix, designations (e.g., PCMH). Also, salient characterizations of the organization, e.g., positioning within the community. \n"
        prompt += "Human-generated summary: \n"
        prompt += "Usually multispecialty clinics, offering variable range of specialties in-house, often including dental care, obstetrics-gynecology, and behavioral health; usually multi-site, occasionally including mobile clinics/vans \n"
        prompt += "FQHCs often play a central role in their communities and healthcare ecosystems \n"
        prompt += "The organization of FQHCs was dynamic, often undergoing expansion \n"
        prompt += "Typical payer mix included Medicaid and Medicare, with varying degrees of commercial insurance and uninsurance. Some FQHCs engaged in value-based contracts with Medicaid and Medicare Advantage. \n"
        prompt += "Several FQHCs in our sample had received PCMH designation or had otherwise been recognized for quality\n\n"
        
        prompt += "Example 2: \n"
        prompt += "Variable: Key contextual facilitators \n"
        prompt += "Variable description: Key positive attributes of the environment, organization, or context, particularly things that could facilitate high quality or high value care for patients with diabetes or other health or social needs \n"
        prompt += "Human-generated summary: \n"
        prompt += "Being deeply embedded in the communities in which they operate, or having personnel from/living in the community \n"
        prompt += "Collaborations with community organizations, e.g., churches, nonprofits, government agencies\n\n"
        
        prompt += "Example 3: \n"
        prompt += "Variable: Community engagement and community-based services \n"
        prompt += "Variable description: Means by which FQHC conducts community needs assessment or how it responds to findings. (Note, patient-specific needs assessment is addressed in the next domain \"support for patient goals, needs, and preferences\") \n"
        prompt += "Means by which the FQHC facilitates connections with community-based entities outside of the FQHC or connections for its patients to these entities; \"means\" could include referral centers, CHWs/promatoras; \"community-based entities\" could include nonprofits, services addressing SDOH \n"
        prompt += "Also, activities the FQHC engages in with community organizations or members outside of the FQHC, in the community.\n\n"
        
        prompt += "Human-generated summary: \n"
        prompt += "Community partnerships and collaborations are often numerous; collaborations fill gaps in care, address health care access and social drivers \n"
        prompt += "Common partners include local hospitals and clinical specialists, schools, state community health center associations, public health departments, social service agencies, nonprofits, religious organizations, courts and correctional facilities, and fire departments \n"
        prompt += "Community-engagement activities encourage patient engagement in health promotion and disease prevention activities"
        
        print("\nFULL PROMPT TO GPT:")
        print("-"*80)
        print(prompt)
        print("-"*80)

        print(f"\nInvoking GPT for variable '{variable}'...")
        comparison_result = gpt.invoke(prompt)

        print("\nGPT RESPONSE:")
        print("-"*80)
        print(comparison_result)
        print("-"*80)

        # Write the comparison result to the new column
        ws.cell(row=row_idx, column=comparison_col_index).value = comparison_result

        # Save the workbook after each row (optional but safer if the process might be long)
        wb.save(output_xlsx)
        print(f"Saved comparison for row {row_idx}")
        print(f"{'-'*60}")

    # Final save after all rows
    wb.save(output_xlsx)
    print(f"[process_spreadsheet()] All done! Saved results to: {output_xlsx}")


if __name__ == "__main__":
    INPUT_XLSX = "site_summaries_all.xlsx"   # Replace with your real file
    OUTPUT_XLSX = "site_summaries_comparison_o1_analysis_730.xlsx"
    process_spreadsheet(INPUT_XLSX, OUTPUT_XLSX)
