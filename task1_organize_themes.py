import time
import requests
import openpyxl
from datetime import datetime
from typing import List, Dict

# ==================================================
# == Stanford GPT Setup (Minimal Example)         ==
# ==================================================
RATE_LIMIT_TOKENS_PER_MINUTE_CHAT = 735000
API_ENDPOINT_CHAT = "https://apim.stanfordhealthcare.org/openai3/deployments/gpt-4o/chat/completions?api-version=2023-05-15"
SUBSCRIPTION_KEY_CHAT = ""

# Track tokens used for chat
tokens_used_chat = 0
start_time_chat = time.time()

class MinimalStanfordSecureGPT:
    def __init__(self, temperature=0.0, max_tokens=25000):
        self.temperature = temperature
        self.max_tokens = max_tokens

    def invoke(self, prompt: str) -> str:
        """
        A basic method to call GPT-4o endpoint with exponential backoff,
        referencing your Stanford secure GPT example.
        """
        global tokens_used_chat, start_time_chat
        max_retries = 5
        retries = 0
        backoff_factor = 2
        initial_wait = 1

        while retries < max_retries:
            # Estimate tokens needed (simplistic approach here)
            tokens_needed = self.max_tokens
            elapsed_time = time.time() - start_time_chat

            # Rate-limiting logic
            if tokens_used_chat + tokens_needed > RATE_LIMIT_TOKENS_PER_MINUTE_CHAT:
                wait_time = 60 - elapsed_time
                if wait_time > 0:
                    print(f"[invoke()] Chat rate limit reached. Waiting {wait_time:.2f} sec...")
                    time.sleep(wait_time)
                tokens_used_chat = 0
                start_time_chat = time.time()

            headers = {
                "Ocp-Apim-Subscription-Key": SUBSCRIPTION_KEY_CHAT,
                "Content-Type": "application/json",
            }
            payload = {
                "messages": [{"role": "user", "content": prompt}],
                "temperature": self.temperature,
                "max_tokens": self.max_tokens,
            }

            try:
                print(f"[invoke()] Sending request to GPT (prompt length={len(prompt)} chars).")
                response = requests.post(API_ENDPOINT_CHAT, headers=headers, json=payload)

                if response.status_code == 429:  # Too Many Requests
                    retry_after = int(response.headers.get("Retry-After", 0))
                    wait_time = retry_after if retry_after > 0 else initial_wait * (backoff_factor ** retries)
                    print(f"[invoke()] 429 Too Many Requests. Sleeping {wait_time}s...")
                    time.sleep(wait_time)
                    retries += 1
                    continue

                response.raise_for_status()
                result = response.json()
                tokens_used_chat += tokens_needed

                answer = result["choices"][0]["message"]["content"]
                print(f"[invoke()] GPT responded with {len(answer)} characters.")
                return answer

            except requests.exceptions.HTTPError as http_err:
                print(f"[invoke()] HTTP error occurred: {http_err}")
                if retries < max_retries - 1:
                    wait_time = initial_wait * (backoff_factor ** retries)
                    print(f"[invoke()] Retrying in {wait_time} sec...")
                    time.sleep(wait_time)
                    retries += 1
                else:
                    raise

            except Exception as e:
                print(f"[invoke()] Error: {e}")
                if retries < max_retries - 1:
                    wait_time = initial_wait * (backoff_factor ** retries)
                    print(f"[invoke()] Retrying in {wait_time} sec...")
                    time.sleep(wait_time)
                    retries += 1
                else:
                    raise

        raise Exception(f"[invoke()] Failed after {max_retries} retries")


def process_spreadsheet(input_xlsx: str, output_xlsx: str):
    """
    1) Load an Excel workbook where:
       - Column A = Variable
       - Column B = Variable description
       - Columns C.. = data from different sites
    2) For each row, build a GPT prompt that compares how each site handles that variable
       based on the variable description.
    3) Write the GPT comparison into a new column at the end.
    4) Save the workbook as we go.
    """

    print(f"[process_spreadsheet()] Loading workbook: {input_xlsx}")
    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb.active  # or choose a specific sheet by name

    # Prepare a GPT instance
    gpt = MinimalStanfordSecureGPT(temperature=0.0, max_tokens=4000)

    # Determine how many columns we currently have
    max_col = ws.max_column
    max_row = ws.max_row

    # Insert a header in the last+1 column for "Comparison across Sites"
    comparison_col_index = max_col + 1
    ws.insert_cols(idx=comparison_col_index)
    ws.cell(row=1, column=comparison_col_index).value = "Comparison"

    # Get site names from the first row
    site_names = {}
    for col_idx in range(4, max_col + 1):
        site_name = ws.cell(row=1, column=col_idx).value
        if site_name:
            site_names[col_idx] = site_name
    
    # Loop through each row, skipping the header row
    for row_idx in range(2, 26):
        variable = ws.cell(row=row_idx, column=1).value
        var_description = ws.cell(row=row_idx, column=2).value
        ccm_definition = ws.cell(row=row_idx, column=3).value

        print(f"\n{'='*80}")
        print(f"Processing Row {row_idx}")
        print(f"Variable: {variable}")
        print(f"Description: {var_description}")
        print(f"Chronic Care Model Definition: {ccm_definition}")
        print('='*80)

        # If variable or description is missing, skip
        if not variable or not var_description:
            print("Skipping row - missing variable or description")
            continue

        # Gather site data from columns C.. up to max_col
        site_data_texts = []
        for col_idx in range(4, max_col + 1):
            site_value = ws.cell(row=row_idx, column=col_idx).value
            if site_value:
                site_name = site_names.get(col_idx, f"Site {col_idx-2}")  # Fallback to Site X if name not found
                site_data_texts.append(f"{site_name}: {site_value}")
                print(f"{site_name} data: {site_value}")

        # If there's no data from any site, skip
        if not site_data_texts:
            print("Skipping row - no site data available")
            continue

        # Build a comparison prompt for GPT
        prompt = (
            f"You are comparing data from different federally qualified health centers.\n\n"
            f"**Variable:** {variable}\n"
            f"**Variable Description:** {var_description}\n\n"
            f"**Chronic Care Model Definition: {ccm_definition}**\n"
            f"Below are notes from different sites:\n"
        )
        for sd in site_data_texts:
            prompt += f"- {sd}\n"

        # prompt += (
        #     "\nPlease compare how this variable is handled or described across the different sites. "
        #     "Identify similarities, differences, or important nuances. "
        #     "Be specific and provide examples."
        #     "Summarize in a clear, concise manner, and list any key insights in bullet points."
        # )
        prompt += (
        "\nPlease identify the main themes "
        "that arise from this data. For each theme, create a structured list of exact quotes from each site "
        "that mention or relate to that theme. Sort the quotes by site. **Do not paraphrase or summarize**; use the direct quotes from the data. "
        "If a site does not mention a particular theme, skip it. "
        "Do not add any analysis or commentary—only group the direct quotes under the relevant theme headings."
        "Create an additional header 'Miscellaneous' and include quotes from data that were not categorized under the themes" 
        "All quotes from the original data should be included in the final output, either under the themes or in the Miscellaneous section, but not in both."
        )


        print("\nFULL PROMPT TO GPT:")
        print("-"*80)
        print(prompt)
        print("-"*80)

        print(f"\nInvoking GPT for variable '{variable}'...")
        comparison_result = gpt.invoke(prompt)

        print("\nGPT RESPONSE:")
        print("-"*80)
        print(comparison_result)
        print("-"*80)

        # Write the comparison result to the new column
        ws.cell(row=row_idx, column=comparison_col_index).value = comparison_result

        # Save the workbook after each row (optional but safer if the process might be long)
        wb.save(output_xlsx)
        print(f"Saved comparison for row {row_idx}")
        print(f"{'-'*60}")

    # Final save after all rows
    wb.save(output_xlsx)
    print(f"[process_spreadsheet()] All done! Saved results to: {output_xlsx}")


if __name__ == "__main__":
    INPUT_XLSX = "site_summaries_all.xlsx"   # Replace with your real file
    OUTPUT_XLSX = "site_summaries_comparison_info_gathering.xlsx"
    process_spreadsheet(INPUT_XLSX, OUTPUT_XLSX)
