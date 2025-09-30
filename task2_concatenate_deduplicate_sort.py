import time
import requests
import re
import openpyxl
from typing import List
from datetime import datetime
from thefuzz import fuzz

# ==============================
# == Rate Limit Setup ==
# ==============================
RATE_LIMIT_TOKENS_PER_MINUTE_CHAT = 25000  # 25k tokens/min
SUBSCRIPTION_KEY_CHAT = ""
API_ENDPOINT_CHAT = "https://apim.stanfordhealthcare.org/openai3/deployments/gpt-4o/chat/completions?api-version=2023-05-15"

tokens_used_chat = 0
start_time_chat = time.time()

def approximate_token_count(text: str) -> int:
    """
    Very rough heuristic: ~4 characters = 1 token.
    This is not exact but helps to stay within your token limit.
    """
    return len(text) // 4

class MinimalStanfordSecureGPT:
    def __init__(self, temperature=0.0, max_tokens=4000):
        self.temperature = temperature
        self.max_tokens = max_tokens

    def invoke(self, prompt: str) -> str:
        """
        Calls the GPT-4 endpoint with exponential backoff,
        while respecting a 25,000 tokens/min rate limit.
        """
        import json
        global tokens_used_chat, start_time_chat
        max_retries = 5
        retries = 0
        backoff_factor = 2
        initial_wait = 1

        print(f"\nStarting invoke() call")
        print(f"Prompt length: {len(prompt)} characters")
        print("num retries: ", retries)
        while retries < max_retries:
            # Approximate total tokens: prompt + potential response
            tokens_needed = approximate_token_count(prompt) + self.max_tokens
            print(f"🔢 Estimated tokens needed: {tokens_needed}")
            
            elapsed_time = time.time() - start_time_chat
            print(f"⏱️ Elapsed time since last reset: {elapsed_time:.2f}s")

            # If exceeding the per-minute limit, wait until reset
            if tokens_used_chat + tokens_needed > RATE_LIMIT_TOKENS_PER_MINUTE_CHAT:
                wait_time = 60 - elapsed_time
                if wait_time > 0:
                    print(f" Chat rate limit reached. Must wait {wait_time:.2f} seconds...")
                    time.sleep(wait_time)

                tokens_used_chat = 0
                start_time_chat = time.time()
                print("🔄 Rate limit counters reset")

            headers = {
                "Ocp-Apim-Subscription-Key": SUBSCRIPTION_KEY_CHAT,
                "Content-Type": "application/json",
            }

            payload = {
                "messages": [{"role": "user", "content": prompt}],
                "temperature": self.temperature,
                "max_tokens": self.max_tokens,
            }

            print(f"Using API key: {SUBSCRIPTION_KEY_CHAT[:8]}...")
            print(f"Sending request to: {API_ENDPOINT_CHAT}")
            
            try:
                print("📤 Sending API request...")
                response = requests.post(
                    API_ENDPOINT_CHAT, headers=headers, json=payload, timeout=60
                )
                
                print(f"Response status code: {response.status_code}")
                print(f"Response headers: {dict(response.headers)}")
                
                if response.status_code != 200:
                    print(f"Error response body: {response.text}")

                if response.status_code == 429:  # Too Many Requests
                    retry_after = int(response.headers.get("Retry-After", 0))
                    wait_time = (
                        retry_after
                        if retry_after > 0
                        else initial_wait * (backoff_factor ** retries)
                    )
                    print(f"429 Too Many Requests. Sleeping {wait_time}s...")
                    time.sleep(wait_time)
                    retries += 1
                    continue

                response.raise_for_status()
                result = response.json()
                print("✅ Successfully got JSON response")

                # Update usage
                tokens_used_chat += tokens_needed
                print(f"📊 Updated tokens used: {tokens_used_chat}")
                
                return result["choices"][0]["message"]["content"]

            except requests.exceptions.HTTPError as http_err:
                print(f"❌ HTTP error occurred: {http_err}")
                print(f"Response body: {response.text}")
                if retries < max_retries - 1:
                    wait_time = initial_wait * (backoff_factor ** retries)
                    print(f"⏳ Retrying in {wait_time} seconds...")
                    time.sleep(wait_time)
                    retries += 1
                else:
                    raise

            except requests.exceptions.Timeout:
                print("⏰ Timeout. Retrying with exponential backoff...")
                retries += 1
                wait_time = initial_wait * (backoff_factor ** retries)
                print(f"⏳ Retrying in {wait_time} seconds...")
                time.sleep(wait_time)
                continue

            except Exception as e:
                print(f"❌ Unexpected error: {str(e)}")
                print(f"Error type: {type(e)}")
                if hasattr(e, '__dict__'):
                    print(f"Error attributes: {e.__dict__}")
                if retries < max_retries - 1:
                    wait_time = initial_wait * (backoff_factor ** retries)
                    print(f"⏳ Retrying in {wait_time} seconds...")
                    time.sleep(wait_time)
                    retries += 1
                else:
                    raise

        raise Exception(f"Failed to get response after {max_retries} retries")
        raise Exception(f"Failed to get response after {max_retries} retries")


def sort_quotes(quotes: List[str], llm: MinimalStanfordSecureGPT, topic: str) -> str:
    """
    Combine the deduplication and sorting steps into one final output.
    """
    print(f"\nStarting sort_quotes for topic: {topic}")
    print(f"Number of quotes to sort: {len(quotes)}")
    
    bullet_points = "".join(f"- {q}\n" for q in quotes)
    print(f"📝 Total length of bullet points: {len(bullet_points)} characters")

    sort_prompt = (
        f"Main Topic: {topic}\n\n"
        "Below is a list of bullet points with quotes.\n"
        "Sort the list based on relevance to the main topic.\n"
        "Instructions:\n"
        "1. Most relevant and important points should come first.\n"
        "2. Maintain the exact format of each bullet point.\n\n"
        f"Bullet Points:\n{bullet_points}\n\n"
        "Sorted List:\n"
    )
    print("sort_prompt: ", sort_prompt)
    print(f"Calling LLM with prompt length: {len(sort_prompt)} characters")
    try:
        result = llm.invoke(sort_prompt)
        print("LLM call successful")
        return result
    except Exception as e:
        print(f"❌ Error in LLM call: {str(e)}")
        print(f"🔍 First 500 chars of prompt: {sort_prompt[:500]}...")
        raise


def parse_bullet_points(text: str) -> List[str]:
    """
    Splits one string containing multiple "[Summary]: ..." sections
    into a list of summary chunks. Each chunk typically includes:
      - The "[Summary]: ..." line
      - One or more [English Quote]: lines.
    We replace Windows line breaks and split on '[Summary]'.
    """
   # print(f"Parsing bullet points: {text}")
    # Convert Windows line breaks
    text = text.replace("\r\n", "\n")
    text = text.replace("•", "")

    # Use a lookahead to split right before each "[Summary]" (ignoring any bullet).
    chunks = re.split(r'(?=\[Summary\])', text.strip())

    # Remove empty or whitespace-only entries
    chunks = [chunk.strip() for chunk in chunks if chunk.strip()]
    filtered_chunks = [c for c in chunks if c.startswith("[Summary]")]

    return filtered_chunks

def remove_duplicate_quotes(bullets: List[str]) -> List[str]:
    """
    Takes a list of bullet chunks. Each chunk might contain one or more
    [English Quote] passages. The code checks for duplicates based on exact
    substring containment (either a new quote contains an old one or vice versa).
    
    Returns a list of bullet chunks, where duplicates have been merged/removed.
    """
    import re

    # pattern = r'\[English Quote\]: "(.*?)"(?=\s*\[DeID[^]]|from \[DeID[^]])'
    #pattern = r'\[English Quote\]: "(.*?)"(?=\s*(?:from \[DeID[^]]|\[DeID[^]]))'
    pattern = r'\[English Quote\]: "(.*?)"(?=\s*(?:from\s*\[[^\]]+|-\s*\[[^\]]+|\[[^\]]+))'

    final_bullets = []
    unique_quotes = []  # parallel to final_bullets

    for bullet in bullets:
        matches = re.findall(pattern, bullet, re.DOTALL)
        
        # If no quotes found, just keep it as-is
        if not matches:
            print(f"No quotes found in bullet: {bullet}")
            continue

    
        new_quote = matches[0]
        # print(f"Quote: {new_quote}, Bullet: {bullet}")
        found_duplicate = False
        for i, existing_quote in enumerate(unique_quotes):
            if existing_quote is None:
                # That bullet had no quotes; skip
                continue

            # Check if new quote is the same or a superset
            if new_quote == existing_quote:
                # Exact match: skip
                #print(f"Exact match: {new_quote} == {existing_quote}")
                found_duplicate = True
                break

            if existing_quote in new_quote:
                # new quote is longer => replace existing
                #print(f"New quote is longer: {new_quote} > {existing_quote}")
                unique_quotes[i] = new_quote
                final_bullets[i] = bullet
                found_duplicate = True
                break

            if new_quote in existing_quote:
                # existing is longer => skip new
                #print(f"New quote is shorter: {new_quote} < {existing_quote}")
                found_duplicate = True
                break

            # do a  fuzzy match 
            score = fuzz.ratio(new_quote, existing_quote)
            if score >= 60:
                # They are considered duplicates
                # print(f"Fuzzy match: {new_quote} ~ {existing_quote} (score={score})")
                found_duplicate = True
                if len(new_quote) > len(existing_quote):
                    unique_quotes[i] = new_quote
                    final_bullets[i] = bullet
                break

        if not found_duplicate:
            # print(f"New quote: {new_quote}")
            unique_quotes.append(new_quote)
            final_bullets.append(bullet)

    return final_bullets
MAX_CHARS_PER_CHUNK = 14000


def chunk_bullets(bullet_points: List[str], max_chars: int = MAX_CHARS_PER_CHUNK) -> List[List[str]]:
    """
    Takes a list of bullet-points. Each bullet can be multiple lines.
    Groups them into sub-lists, each up to 'max_chars' in total length.
    Ensures that we only split on bullet boundaries, so we never cut a bullet in half.
    """
    chunks = []
    current_chunk = []
    current_length = 0

    for i in range(len(bullet_points)):
        if i != 0:
            continue
            #i want to find the word "Eddie" in the bullet point and change it to "eddie"
        if i == 0:
            bullet = bullet_points[i]
            # i want it to remove all the parts after [Spanish Quote]
            # i want it to find and replace "por decirte" from the quote
            bullet = bullet.replace("por decirte", "")



            bullet_points[i] = bullet
            print(f"first bullet: {bullet_points[i]}")
            
        bullet = bullet_points[i]
        bullet_len = len(bullet)
        # If adding this bullet would exceed max_chars, start a new chunk
        if current_length + bullet_len > max_chars and current_chunk:
            chunks.append(current_chunk)
            current_chunk = [bullet]
            current_length = bullet_len
        else:
            current_chunk.append(bullet)
            current_length += bullet_len

    # Add the last chunk if not empty
    if current_chunk:
        chunks.append(current_chunk)

    return chunks

def process_dedup_and_sort(deduped_answers: List[str], llm, topic: str) -> str:
    total_char_count = sum(len(b) for b in deduped_answers)
    print(f"\nProcessing dedup and sort for topic: {topic}")
    print(f"Total characters: {total_char_count}")

    if total_char_count > MAX_CHARS_PER_CHUNK:
        print(f"Content exceeds {MAX_CHARS_PER_CHUNK} chars, splitting into chunks")
        bullet_chunks = chunk_bullets(deduped_answers, max_chars=MAX_CHARS_PER_CHUNK)
        print(f"Split into {len(bullet_chunks)} chunks")

        sorted_chunks = []
        for i, chunk in enumerate(bullet_chunks):
           
            print(f"\nProcessing chunk {i+1}/{len(bullet_chunks)}")
            print(f"Chunk size: {sum(len(b) for b in chunk)} characters")
            try:
                sorted_chunk_text = sort_quotes(chunk, llm, topic)
                sorted_chunks.append(sorted_chunk_text)
                print(f"✅ Successfully processed chunk {i+1}")
            except Exception as e:
                print(f"❌ Error processing chunk {i+1}: {str(e)}")
                raise

        final_result = "\n\n".join(sorted_chunks)
        print("All chunks processed and combined")
    else:
        print("Content within limits, processing as single chunk")
        final_result = sort_quotes(deduped_answers, llm, topic)
        print("Processing complete")

    return final_result


def process_spreadsheet(input_xlsx: str, output_xlsx: str):
    """
    Load 'input_xlsx', but output everything to a single sheet in a brand-new workbook:
      - The new workbook has exactly one sheet named 'Consolidated'.
      - Each column in 'Consolidated' is labeled for one input sheet (sheet_1, sheet_2, etc.).
      - For row r in the original sheet, deduplicate and sort its sub-answers,
        then put the result in row r, in the column belonging to that sheet.
      - The first column (A) is used for site_ids. 
    """
    print(f"[process_spreadsheet()] Loading {input_xlsx}")
    wb_in = openpyxl.load_workbook(input_xlsx)

    # Create a brand-new workbook with a single sheet
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Consolidated"

    sheet_names = wb_in.sheetnames
    print("Input sheets:", sheet_names)

    # We'll name the columns in row=1:
    # Col A => "site_id"
    # Col B => sheet_names[0]
    # Col C => sheet_names[1]
    # ...
    ws_out.cell(row=1, column=1, value="site_id")
    for i, sname in enumerate(sheet_names, start=2):
        ws_out.cell(row=1, column=i, value=sname)

    llm = MinimalStanfordSecureGPT()

    # Process each sheet one by one
    for sheet_index, sheet_name in enumerate(sheet_names):
        ws_in = wb_in[sheet_name]
        max_col = ws_in.max_column
        max_row = ws_in.max_row

        print(f"\n=== PROCESSING SHEET: {sheet_name}")
        print(f"    Rows={max_row}, Cols={max_col}")

        # The column in the new sheet for this sheet
        out_col = sheet_index + 2  # sheet_index 0 => col 2, index 1 => col 3, etc.

        # For each row in the input sheet, gather sub-answers, run LLM, store results
        for row_i in range(2, max_row + 1):
            site_id = ws_in.cell(row=row_i, column=1).value
            if not site_id:
                continue

            # Skip if not S_02 or S_03
            # if not (site_id.startswith('S_02') or site_id.startswith('S_03')):
                
            #     continue

            if not (site_id.startswith('S_01')):
                
                continue

            print(f"\nProcessing site_id: {site_id}")

            # In the new sheet, row = row_i is the same row as input
            if not ws_out.cell(row=row_i, column=1).value:
                ws_out.cell(row=row_i, column=1, value=str(site_id))

            # Gather sub-answers from columns 2..max_col
            sub_answers = []
            for c in range(2, max_col + 1):
                cell_val = ws_in.cell(row=row_i, column=c).value
                if cell_val and isinstance(cell_val, str):
                    bullet_points = parse_bullet_points(cell_val.strip())
                    sub_answers.extend(bullet_points)

            if not sub_answers:
                continue

            deduped_answers = remove_duplicate_quotes(sub_answers)
            sorted_answers = process_dedup_and_sort(deduped_answers, llm, sheet_name)

            if len(sorted_answers) > 32767:
                print(f"WARNING: sorted_answers for {site_id} in {sheet_name} is too long: {len(sorted_answers)} chars")

            ws_out.cell(row=row_i, column=out_col, value=sorted_answers)
            wb_out.save(output_xlsx)

    wb_out.save(output_xlsx)
    print(f"[All done. Saved => {output_xlsx}]")


# Example usage:
if __name__ == "__main__":
    # INPUT_XLSX = "all_questions_final_including_negatives.xlsx"
    # OUTPUT_XLSX = "all_questions_analysis_manual_dedup_7.xlsx"

    INPUT_XLSX = "cpts_s.xlsx"
    OUTPUT_XLSX = "all_questions_analysis_manual_dedup_14000.xlsx"

    process_spreadsheet(INPUT_XLSX, OUTPUT_XLSX) 


